import pandas as pd
import numpy as np
from sqlalchemy.orm import Session
from app.db.models import ValidationError
import os
import openpyxl
from openpyxl.styles import PatternFill
import re

def is_valid_email(email):
    if pd.isna(email): return True
    return re.match(r"[^@]+@[^@]+\.[^@]+", str(email)) is not None

def validate_excel(file_path: str, upload_id: int, db: Session, report_path: str):
    try:
        df = pd.read_excel(file_path)
    except Exception as e:
        error = ValidationError(upload_id=upload_id, error_message=f"Could not read Excel file: {str(e)}")
        db.add(error)
        db.commit()
        return 1

    errors = []
    
    # Required columns
    required_columns = ['Name', 'Email', 'Amount', 'Date']
    missing_cols = [col for col in required_columns if col not in df.columns]
    for col in missing_cols:
        errors.append(ValidationError(upload_id=upload_id, error_message=f"Missing required column: {col}"))
        
    if missing_cols:
        for err in errors:
            db.add(err)
        db.commit()
        return len(errors)

    for index, row in df.iterrows():
        row_num = index + 2 
        
        if pd.isna(row.get('Name')):
            errors.append(ValidationError(upload_id=upload_id, row_number=row_num, column_name='Name', error_message="Name is empty"))
            
        email = row.get('Email')
        if not pd.isna(email) and not is_valid_email(email):
            errors.append(ValidationError(upload_id=upload_id, row_number=row_num, column_name='Email', error_message="Invalid email format"))
            
        amount = row.get('Amount')
        if pd.isna(amount):
             errors.append(ValidationError(upload_id=upload_id, row_number=row_num, column_name='Amount', error_message="Amount is empty"))
        elif not isinstance(amount, (int, float, np.number)) or amount <= 0:
             errors.append(ValidationError(upload_id=upload_id, row_number=row_num, column_name='Amount', error_message="Amount must be > 0"))

    for err in errors:
        db.add(err)
    db.commit()

    if errors:
        generate_error_report(file_path, report_path, errors)
    
    return len(errors)

def generate_error_report(input_file: str, output_file: str, errors: list):
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active
    
    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    
    col_map = {cell.value: i+1 for i, cell in enumerate(ws[1])}
    
    for err in errors:
        if err.row_number and err.column_name in col_map:
            col_idx = col_map[err.column_name]
            ws.cell(row=err.row_number, column=col_idx).fill = red_fill
            
    summary_ws = wb.create_sheet("Error Summary")
    summary_ws.append(["Row", "Column", "Error Message"])
    for err in errors:
        summary_ws.append([err.row_number, err.column_name, err.error_message])
        
    wb.save(output_file)
